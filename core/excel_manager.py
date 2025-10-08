"""Excel 기반 번역 검수 워크플로우

원문 + 번역 → Excel 내보내기 → 수정 → Excel 업로드 → 스마트 병합
"""
import hashlib
import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import json


class TranslationEntry:
    """번역 항목"""

    def __init__(
        self,
        file_path: str,
        line_number: int,
        japanese: str,
        translation: str,
        entry_id: Optional[str] = None
    ):
        self.file_path = file_path
        self.line_number = line_number
        self.japanese = japanese
        self.translation = translation
        self.entry_id = entry_id or self._generate_id()

        self.modified_translation: Optional[str] = None
        self.status = "pending"  # pending, reviewed, modified
        self.reviewer: Optional[str] = None
        self.review_date: Optional[str] = None

    def _generate_id(self) -> str:
        """고유 ID 생성 (충돌 방지용)"""
        # 파일명 + 줄번호 + 원문 해시 (경로 제외, 파일명만 사용)
        from pathlib import Path
        filename = Path(self.file_path).name
        raw = f"{filename}:{self.line_number}:{self.japanese}"
        return hashlib.md5(raw.encode('utf-8')).hexdigest()[:12]

    def to_dict(self) -> dict:
        """딕셔너리 변환"""
        return {
            'ID': self.entry_id,
            '파일명': Path(self.file_path).name,
            '줄번호': self.line_number,
            '원문': self.japanese,
            'AI 번역': self.translation,
            '수정본': self.modified_translation or ''
        }

    @classmethod
    def from_dict(cls, data: dict) -> 'TranslationEntry':
        """딕셔너리로부터 생성"""
        entry = cls(
            file_path=data.get('파일명', ''),
            line_number=data.get('줄번호', 0),
            japanese=data.get('원문', ''),
            translation=data.get('AI 번역', ''),
            entry_id=data.get('ID')
        )
        entry.modified_translation = data.get('수정본') or None
        # 수정된 번역이 있으면 상태를 modified로 변경
        if entry.modified_translation:
            entry.status = 'modified'
        return entry


class ExcelManager:
    """Excel 번역 관리자"""

    def export_to_excel(
        self,
        entries: List[TranslationEntry],
        output_path: str,
        modified_only: bool = False
    ):
        """Excel로 내보내기

        Args:
            entries: 번역 항목 리스트
            output_path: 출력 Excel 파일 경로
            modified_only: 수정된 항목만 내보내기
        """
        # 필터링
        if modified_only:
            entries = [e for e in entries if e.status == 'modified']

        # DataFrame 생성 (dict와 TranslationEntry 모두 처리)
        data = []
        for i, e in enumerate(entries):
            if isinstance(e, dict):
                # Unity 게임용 dict
                data.append({
                    'ID': i + 1,
                    '파일명': e.get('file', 'unknown'),
                    '줄번호': '',  # Unity 게임은 줄번호 없음
                    '원문': e.get('original', ''),
                    'AI 번역': e.get('translated', ''),
                    '수정본': ''
                })
            else:
                # Naninovel용 TranslationEntry 객체
                data.append(e.to_dict())

        df = pd.DataFrame(data)

        # Excel 저장 (스타일 적용)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='번역')

            # 워크시트 가져오기
            worksheet = writer.sheets['번역']

            # 열 너비 자동 조정
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 80)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 헤더 스타일
            from openpyxl.styles import Font, PatternFill, Alignment

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # '수정본' 열 강조 (노란색)
            modified_col = df.columns.get_loc('수정본') + 1
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in range(2, len(df) + 2):
                worksheet.cell(row=row, column=modified_col).fill = yellow_fill

        print(f"✅ Excel 내보내기 완료: {output_path}")
        print(f"   항목 수: {len(entries)}개")

    def import_from_excel(
        self,
        excel_path: str,
        original_entries: List[TranslationEntry]
    ) -> Tuple[List[TranslationEntry], List[Dict]]:
        """Excel에서 가져오기 및 병합

        Args:
            excel_path: Excel 파일 경로
            original_entries: 원본 번역 항목

        Returns:
            (병합된 항목 리스트, 충돌 리스트)
        """
        # Excel 읽기
        df = pd.read_excel(excel_path)

        # ID 기반 매핑
        original_map = {e.entry_id: e for e in original_entries}
        conflicts = []
        updated_count = 0

        print(f"📊 Excel 컬럼: {list(df.columns)}")
        print(f"📊 Excel 총 행 수: {len(df)}")

        empty_count = 0
        same_count = 0

        for idx, row in df.iterrows():
            entry_id = row['ID']
            modified = str(row.get('수정본', '')).strip()

            # 수정본이 비어있으면 건너뛰기
            if not modified or modified == 'nan':
                empty_count += 1
                continue

            if entry_id in original_map:
                original = original_map[entry_id]

                # 수정 사항 반영 (수정된 내용이 AI 번역과 다를 때만)
                if modified != original.translation:
                    original.modified_translation = modified
                    original.status = 'modified'
                    updated_count += 1
                    if updated_count <= 3:  # 처음 3개만 출력
                        print(f"  ✏️ 수정: {original.translation[:30]}... → {modified[:30]}...")
                else:
                    same_count += 1
            else:
                # ID 없음 → 충돌
                conflicts.append({
                    'id': entry_id,
                    'file': row.get('파일명'),
                    'line': row.get('줄번호'),
                    'reason': 'ID not found in original'
                })
                if len(conflicts) <= 3:  # 처음 3개만 출력
                    print(f"  ⚠️ ID 없음: {entry_id}")

        print(f"✅ Excel 가져오기 완료:")
        print(f"   수정 항목: {updated_count}개")
        print(f"   빈 수정본: {empty_count}개")
        print(f"   AI 번역과 동일: {same_count}개")
        print(f"   충돌: {len(conflicts)}개")

        return list(original_map.values()), conflicts

    def apply_to_files(
        self,
        entries: List[TranslationEntry],
        output_dir: Path
    ):
        """번역 파일에 적용

        Args:
            entries: 번역 항목 (수정된 번역 포함)
            output_dir: 출력 디렉토리
        """
        # 파일별로 그룹화
        file_groups = {}
        for entry in entries:
            file_path = entry.file_path
            if file_path not in file_groups:
                file_groups[file_path] = []
            file_groups[file_path].append(entry)

        # 파일별 처리
        for file_path, file_entries in file_groups.items():
            # 원본 파일 읽기
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            # 수정 사항 적용 (줄번호 기반)
            for entry in file_entries:
                if entry.modified_translation:
                    line_idx = entry.line_number - 1  # 0-based index
                    if 0 <= line_idx < len(lines):
                        # 번역 라인 교체
                        lines[line_idx] = entry.modified_translation + '\n'

            # 출력 파일 저장
            output_path = output_dir / Path(file_path).name
            with open(output_path, 'w', encoding='utf-8') as f:
                f.writelines(lines)

        print(f"✅ {len(file_groups)}개 파일에 번역 적용 완료")

    def save_history(
        self,
        entries: List[TranslationEntry],
        history_path: str
    ):
        """번역 이력 저장

        Args:
            entries: 번역 항목
            history_path: 이력 JSON 파일 경로
        """
        history_data = {
            'timestamp': datetime.now().isoformat(),
            'entries': [e.to_dict() for e in entries if e.status == 'modified']
        }

        with open(history_path, 'w', encoding='utf-8') as f:
            json.dump(history_data, f, ensure_ascii=False, indent=2)

        print(f"✅ 번역 이력 저장: {history_path}")


# 사용 예시
if __name__ == "__main__":
    # 1. 번역 항목 생성
    entries = [
        TranslationEntry("test.txt", 10, "こんにちは", "안녕하세요"),
        TranslationEntry("test.txt", 20, "ありがとう", "감사합니다"),
    ]

    # 2. Excel 내보내기
    manager = ExcelManager()
    manager.export_to_excel(entries, "translation_review.xlsx")

    # 3. Excel 수정 후 가져오기
    # updated_entries, conflicts = manager.import_from_excel("translation_review.xlsx", entries)

    # 4. 파일에 적용
    # manager.apply_to_files(updated_entries, Path("output"))
